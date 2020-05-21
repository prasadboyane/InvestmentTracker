import openpyxl
import os.path
import time
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.pyplot import figure

from pass_recovery import *


import base64
import os
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.fernet import Fernet



def create_user():
    #Create User
    user=input('Please Enter New Username: ')
    password=input('Please Enter New Password: ')
    confirm_password=input('Please confiem Password again: ')
    create_dtm = time.strftime("%x")
    mobile_number = input('Please enter your 10 digit Mobile Number: ')
    if len(mobile_number) !=10:
        print('Please enter valid Mobile Number as it is required for password recovery !')
        create_user()
    
    if password == confirm_password:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Metadata"
        wb.save('Tracker_enc.xlsx')
        
        
        sheet['A1'] = user
        sheet['B1'] = password
        sheet['C1'] = create_dtm
        sheet['D1'] = mobile_number
        
        
        sheet1 = wb.create_sheet("data")
        sheet1 = wb["data"]
        sheet1['A1'] = 'id'
        sheet1['B1'] = 'amount'
        sheet1['C1'] = 'date'
        sheet1['D1'] = 'assetname'
        sheet1['E1'] = 'Maturity date'
        wb.save('Tracker_enc.xlsx')
        """
        Generates a key and save it into a file
        """
        key = Fernet.generate_key()
        with open("key.key", "wb") as key_file:
            key_file.write(key)        
            
        
        encrypt_file('Tracker_enc.xlsx',load_key())
        
    else:
        print()
        print('Passwords did not match. Please try once again !')
        print()
        create_user()
        

def generate_Pcode():
    import random
    import math
    
    digits = [i for i in range(0, 10)]
    
    random_pcode = ""
    
    for i in range(6):
        index = math.floor(random.random() * 10)
        random_pcode += str(digits[index])
    
    return random_pcode

def recover_pass():
    mob_no=input('Enter your registered 10 digit Mobile Number:')
    if len(mob_no) != 10:
        print('Please enter valid 10 digit Mobile number')
        recover_pass()
    
    decrypt_file('Tracker_enc.xlsx',load_key())
    wb = openpyxl.load_workbook('Tracker_enc.xlsx')
    ws = wb["Metadata"]
    reg_mob_no = ws['D1'].value
    encrypt_file('Tracker_enc.xlsx',load_key())
    
    if mob_no != reg_mob_no:
        print('This mobile number is not registered ! ')
        ip=input('press r to retry or e to exit: ')
        if ip =='r':
            recover_pass()
        else:
            exit()
    
    else:
        pcode = generate_Pcode()
        response = sendPostRequest(URL, '<provided-api-key>', '<secret-api-key>', 'stage', reg_mob_no, '<active-sender-id>', 'Hi, Your P-CODE is: {}. Please do not share it with anyone'.format(pcode) )
        if '"status":"success"' in response.text:
            print('P-CODE sent successfully !! ')
            rc_pcode=input('Please enter 6 digit P-CODE: ')
            if rc_pcode == pcode:
                new_pass=input('Please enter new password: ')
                decrypt_file('Tracker_enc.xlsx',load_key())
                wb = openpyxl.load_workbook('Tracker_enc.xlsx')
                ws = wb["Metadata"]
                ws['B1'] = new_pass
                wb.save('Tracker_enc.xlsx') 
                encrypt_file('Tracker_enc.xlsx',load_key())
                login_user()
        else:
            print('Something went wrong. Please try again.')
            recover_pass()

def login_user():
    #Create User
    user=input('Please Enter Username: ')
    password=input('Please Enter Password: ')
    
    decrypt_file('Tracker_enc.xlsx',load_key())
    wb = openpyxl.load_workbook('Tracker_enc.xlsx')
    ws = wb["Metadata"]   
    if user == ws['A1'].value  and password == ws['B1'].value:
        encrypt_file('Tracker_enc.xlsx',load_key())
        print('logged in !')
    else:
        encrypt_file('Tracker_enc.xlsx',load_key())
        print('password is incorrect ! Try again !')
        print('****     1. Try again        *******')
        print('****     2. Forgot Password? *******')
        print('****     3. exit *******')
        retry_creds = input('Please Enter your choice: ')
        if retry_creds == '2':
            recover_pass()
        elif retry_creds == '3':
            exit()
        else:
            login_user()
    
    

def insert_entry():
    amount_entry = input('Please enter amount: ')
    date = input('Enter a date (YYYY-MM-DD)')
    try:
        year, month, day = map(int, date.split('-'))
        date_entry = datetime(year, month, day)
    except:
        print('Please valid enter date in correct format !')
        insert_entry()
    
    print('----------------')
    print('1. FD')
    print('2. RD')
    print('3. MF')
    print('----------------')
    inv_type_entry=input('Select Category Number: ')
    if not (inv_type_entry == '1' or inv_type_entry == '2' or inv_type_entry == '3'):
        print('Please enter select category number !')
        insert_entry()
    
    wb = openpyxl.load_workbook('Tracker_enc.xlsx')
    data_sheet = wb['data']
    
    try:
        max_row_for_c = max((c.row for c in data_sheet['A'] if c.value is not None))  # To find max number of rows in 'A' columns
    except:
        max_row_for_c=0
        
    row = (max_row_for_c,amount_entry,str(date_entry),inv_type_entry)
    data_sheet.append(row)
    wb.save('Tracker_enc.xlsx') 
    go_next()

def update_entry():
    id_entry = input('Please enter id to update: ')
    id_entry_int=int(id_entry)+1
    id_entry_1=str(id_entry_int)
    
    wb = openpyxl.load_workbook('Tracker_enc.xlsx')
    data_sheet = wb["data"]
    i_a = 'A{}'.format(id_entry_1)
    i_b = 'B{}'.format(id_entry_1)
    i_c = 'C{}'.format(id_entry_1)
    i_d = 'D{}'.format(id_entry_1)
    i_e = 'E{}'.format(id_entry_1)
    
    msg="Here's your record: ",data_sheet[i_a].value,data_sheet[i_b].value,data_sheet[i_c].value,data_sheet[i_d].value,data_sheet[i_e].value
    print(msg)
    print('which field you want to edit?')
    print('1. amount')
    print('2. date')
    print('3. assetname')
    print('4. maturity date')
    print('5. Back')
    choice=input('Enter choice number: ')
    
    if choice=='1':
        updated_val=input('Enter updated value for amount: ')
        data_sheet[i_b] = updated_val
    elif choice =='2':
        updated_val=input('Enter updated value for date(YYYY-MM-DD): ')
        try:
            year, month, day = map(int, updated_val.split('-'))
            updated_val_1 = datetime(year, month, day)
        except:
            print('Please valid enter date in correct format !')
            update_entry()
        data_sheet[i_c] = updated_val_1
    elif choice =='3':
        updated_val=input('Enter updated value for assetname: ')
        data_sheet[i_d] = updated_val
    elif choice =='4':
        updated_val=input('Enter updated value for Maturity date(YYYY-MM-DD): ')
        try:
            year, month, day = map(int, updated_val.split('-'))
            updated_val_1 = datetime(year, month, day)
        except:
            print('Please valid enter date in correct format !')
            update_entry()
        data_sheet[i_c] = updated_val_1
    elif choice =='5':
        show_menu()
        
    wb.save('Tracker_enc.xlsx') 
    go_next()


def view_entry():
    wb = openpyxl.load_workbook('Tracker_enc.xlsx')
    ws = wb["data"]
    max_row_for_a = max((c.row for c in ws['A'] if c.value is not None))
    
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=max_row_for_a, max_col=5):
        for cell in row:
            print(cell.value,'\t\t', end=" ")
        print()
    go_next()


def delete_entry():
    wb = openpyxl.load_workbook('Tracker_enc.xlsx')
    ws = wb["data"]
    
    id_entry = input('Please enter id to delete: ')
    ws.delete_rows(int(id_entry)+1)
    wb.save('Tracker_enc.xlsx')
    go_next()
    

def view_graph():
    
    
    fig = plt.figure(num=None, figsize=(13, 6), dpi=80, facecolor='w', edgecolor='k')
    man = plt.get_current_fig_manager()
    man.canvas.set_window_title("Investment Dashboard")
    
    wb = openpyxl.load_workbook('Tracker_enc.xlsx')
    ws = wb["data"]
    max_row_for_a = max((c.row for c in ws['A'] if c.value is not None))
    
    #To set X axis based on dates
    date_arr=[]
    colC = ws['C']
    for cell in colC:
        if cell.value != 'date':
            date_arr.append(cell.value)
    date_arr.sort()
    min_date=min(date_arr)
    max_date=max(date_arr)
    
    
    #To set total amounts Y axis
    tot_amount_arr=[]
    colB = ws['B']
    for cell in colB:
        if cell.value != 'amount':
            tot_amount_arr.append(int(cell.value))  
    tot_amount_arr.sort()
    min_am=str(min(tot_amount_arr))
    max_am=str(max(tot_amount_arr))
    total_inv=sum(tot_amount_arr)
  
  
    #generating 3 y-amount axis
    fd_y_list=[]
    rd_y_list=[]
    mf_y_list=[]
   
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=4, max_row=max_row_for_a):
        for cell in row:
            if row[2].value == '1':
                fd_y_list.append(cell.value)
            elif row[2].value == '2':
                rd_y_list.append(cell.value)
            else:
                mf_y_list.append(cell.value)
                
    fd_y_list=list(filter(lambda a: a != '1', fd_y_list))
    rd_y_list=list(filter(lambda a: a != '2', rd_y_list))
    mf_y_list=list(filter(lambda a: a != '3', mf_y_list))
    
    fd_y_list =[ x for x in fd_y_list if "-" not in x ]
    rd_y_list =[ x for x in rd_y_list if "-" not in x ]
    mf_y_list =[ x for x in mf_y_list if "-" not in x ]
    
    #convert to int as matplot y-axis array should be numeric to sort properly
    fd_y_list=[int(i) for i in fd_y_list] 
    rd_y_list=[int(i) for i in rd_y_list] 
    mf_y_list=[int(i) for i in mf_y_list] 
    
    #print('y axis: ',fd_y_list)
    #print('y axis: ',rd_y_list)
    #print('y axis: ',mf_y_list) 
   
    #generating 3 x-amount axis
    fd_x_list=[]
    rd_x_list=[]
    mf_x_list=[]
   
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4, max_row=max_row_for_a):
        for cell in row:
            if row[1].value == '1':
                fd_x_list.append(cell.value)
            elif row[1].value == '2':
                rd_x_list.append(cell.value)
            else:
                mf_x_list.append(cell.value)
                
    fd_x_list=list(filter(lambda a: a != '1', fd_x_list))
    rd_x_list=list(filter(lambda a: a != '2', rd_x_list))
    mf_x_list=list(filter(lambda a: a != '3', mf_x_list))
    
    fd_x_list=[datetime.strptime(i, '%Y-%m-%d %H:%M:%S') for i in fd_x_list] 
    rd_x_list=[datetime.strptime(i, '%Y-%m-%d %H:%M:%S') for i in rd_x_list] 
    mf_x_list=[datetime.strptime(i, '%Y-%m-%d %H:%M:%S') for i in mf_x_list] 
    #datee2 = datetime.strptime(datee, '%Y-%m-%d %H:%M:%S')
    
    #print('x axis: ',fd_x_list)
    #print('x axis: ',rd_x_list)
    #print('x axis: ',mf_x_list)  

            
                
    plt.plot(fd_x_list, fd_y_list, color='#03DAC6', linestyle='-', marker='o',linewidth=0.8, label='FD')
    plt.plot(rd_x_list, rd_y_list, color='#018786', linestyle='-', marker='o',linewidth=0.8, label='RD')
    plt.plot(mf_x_list, mf_y_list, color='#f44336', linestyle='-', marker='o',linewidth=0.8, label='MF')
    
    #plt.axis([min_date, max_date, '0', max_am])
    plt.title('Your Investment Journey: {} INR'.format(total_inv))
    #plt.text(1,1, r'$Total Investment={}$'.format(total_inv), bbox=dict(facecolor='blue', alpha=0.5))
    plt.ylabel('Amount (INR)')
    plt.xlabel('Tenure (Months)')
    
    plt.legend() #to tag each graph to its label
    #plt.tight_layout() # To adjust as per screen size
    plt.show()
    
    go_next()
    
def go_next():
    nxt = input('Press "m" for Main Menu & "e" for exit ')
    if nxt=='m':
        show_menu()
    elif nxt=='e':
        exit()
    else:
        print('Please enter valid choice !')
        go_next()
     


def load_key():
    return open("key.key", "rb").read()

def encrypt_file(filename, key):
    """
    Given a filename (str) and key (bytes), it encrypts the file and write it
    """
    #print('encrypting file...')
    f = Fernet(key)
    
    with open(filename, "rb") as file:
        # read all file data
        file_data = file.read()
        
    encrypted_data = f.encrypt(file_data)
    
    with open(filename, "wb") as file:
        file.write(encrypted_data)    
        


def decrypt_file(filename, key):
    """
    Given a filename (str) and key (bytes), it decrypts the file and write it
    """
    #print('decrypting file...')
    f = Fernet(key)
    with open(filename, "rb") as file:
        # read the encrypted data
        encrypted_data = file.read()
    # decrypt data
    decrypted_data = f.decrypt(encrypted_data)
    # write the original file
    with open(filename, "wb") as file:
        file.write(decrypted_data)


def show_menu():
    print('******   1. Insert new Entry ******')
    print('******   2. Update Entry     ******')
    print('******   3. Delete Entry     ******')
    print('******   4. View Entries     ******')
    print('******   5. View Graph       ******')
    print('******   6. Exit             ******')
    choice = input('Please enter your choice: ')
    if choice == '1':
        insert_entry()
    elif choice == '2':
        update_entry()
    elif choice == '3':
        delete_entry()
    elif choice == '4':
        view_entry()
    elif choice == '5':
        view_graph()
    elif choice == '6':
        pass
    else:
        print('Enter valid choice !')
        show_menu()
        

##START PROGRAM HERE
try:
    if os.path.exists('Tracker_enc.xlsx') == True:
        login_user()
    else:
        create_user()
    
    decrypt_file('Tracker_enc.xlsx',load_key())
    show_menu()
except Exception as e:
    print('Something went wrong ! ERROR: ', e)

finally:
    encrypt_file('Tracker_enc.xlsx',load_key())
    exit()  
